
import re
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def _celdas_a_dataframe(ws, rango_celdas):
    """
    Convierte un rango de celdas de una hoja de cálculo de openpyxl a un DataFrame de pandas.
    """
    datos = []
    # Divide el rango en celdas de inicio y fin para iterar
    min_col, min_row, max_col, max_row = rango_celdas
    
    # Itera sobre las filas y columnas del rango para obtener los datos
    for fila in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True):
        datos.append(fila)
    
    # Crea un DataFrame. Si hay datos, usa la primera fila como encabezados.
    if not datos:
        return pd.DataFrame()
    
    # Nombres de columna por defecto si no hay encabezados
    column_names = [get_column_letter(c) for c in range(min_col, max_col + 1)]
    df = pd.DataFrame(datos, columns=column_names)
    
    return df

from collections import defaultdict

def extraer_bloques_desde_hoja(ws, formatos_config):
    """
    Analiza una hoja de cálculo de openpyxl en busca de códigos especiales y extrae
    bloques de contenido (texto y tablas) en el orden en que aparecen.
    """
    bloques_con_posicion = []
    # Usar defaultdict(list) para almacenar múltiples posiciones para el mismo ID de tabla
    pos_inicio_tablas = defaultdict(list)
    pos_fin_tablas = defaultdict(list)
    
    patron_codigo = re.compile(r'\[\[(.*?)\]\]')

    # Antes de iterar, verificar si la hoja tiene alguna dimensión.
    # Si no tiene max_row o max_column, es probable que esté vacía.
    if not ws.max_row or not ws.max_column:
        return []

    # Se vuelve a `iter_rows` por rendimiento, pero con un manejo de errores
    # más robusto para capturar el `KeyError` si ocurre.
    try:
        for fila in ws.iter_rows():
            for celda in fila:
                if isinstance(celda.value, str):
                    match = patron_codigo.search(celda.value)
                    if match:
                        codigo_completo = match.group(1).strip()
                        
                        if codigo_completo.startswith('inicio_'):
                            id_tabla = codigo_completo.replace('inicio_', '')
                            pos_inicio_tablas[id_tabla].append((celda.row, celda.column + 1))
                        
                        elif codigo_completo.startswith('fin_'):
                            id_tabla = codigo_completo.replace('fin_', '')
                            pos_fin_tablas[id_tabla].append((celda.row, celda.column - 1))

                        else:
                            id_simple = codigo_completo
                            if id_simple in formatos_config.get('tipos', {}):
                                # El contenido está en la celda de al lado
                                contenido_celda = ws.cell(row=celda.row, column=celda.column + 1).value
                                bloques_con_posicion.append({
                                    'tipo': id_simple,
                                    'contenido': contenido_celda or "",
                                    'fila': celda.row
                                })
    except KeyError as e:
        print(f"Advertencia: Ocurrió un error de clave al procesar la hoja '{ws.title}'. "
              f"Esto puede suceder con hojas anómalas. La hoja será omitida. Error: {e}")
        return []
    except Exception as e:
        print(f"Error inesperado procesando la hoja '{ws.title}': {e}")
        return []

    # 2. Consolidar las tablas en la lista de bloques
    for id_tabla, inicios in pos_inicio_tablas.items():
        fines = pos_fin_tablas.get(id_tabla, [])
        
        # Asegurarse de que haya el mismo número de inicios y fines
        if len(inicios) != len(fines):
            # Podríamos añadir una advertencia aquí si es necesario
            continue
        
        # Ordenar ambas listas por fila para asegurar el emparejamiento correcto
        inicios.sort()
        fines.sort()
        
        for i in range(len(inicios)):
            pos_inicio = inicios[i]
            pos_fin = fines[i]
            
            # Comprobar que el fin está después del inicio
            if pos_fin[0] < pos_inicio[0]:
                continue

            rango = (pos_inicio[1], pos_inicio[0], pos_fin[1], pos_fin[0])
            
            df_tabla = _celdas_a_dataframe(ws, rango)
            
            if not df_tabla.empty:
                # Añadir el bloque de tabla con su fila de inicio
                bloques_con_posicion.append({
                    'tipo': id_tabla,
                    'contenido': df_tabla,
                    'fila': pos_inicio[0] # Fila de inicio de la tabla
                })

    # 3. Ordenar la lista completa de bloques por su número de fila
    bloques_ordenados = sorted(bloques_con_posicion, key=lambda b: b['fila'])

    # 4. Limpiar la clave 'fila' que ya no es necesaria
    for bloque in bloques_ordenados:
        del bloque['fila']

    return bloques_ordenados
