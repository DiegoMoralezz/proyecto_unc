# scripts/motor_automatizacion.py
from __future__ import annotations

from pathlib import Path
import json
import sys
from io import BytesIO
import gc  # Importar el módulo de recolección de basura

# Dependencias de procesamiento
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

# --- 1. ÚNICA FUENTE DE VERDAD PARA RUTAS ---
# Se definen rutas relativas desde la raíz del proyecto. Streamlit Cloud
# ejecuta la app desde la raíz del repo, por lo que estas rutas son portables.
CONFIG_PATH = Path("config/rangos_hojas.json")
FORMATOS_PATH = Path("config/formatos_hojas.json")
PLANTILLA_PATH = Path("plantilla/plantilla_base_final.docx")


# --- 2. GESTIÓN DE IMPORTS INTERNOS ---
# Se elimina la manipulación de sys.path. Al ejecutar `streamlit run app.py`
# desde la raíz, Python maneja los módulos en `scripts/` correctamente.
from scripts.core_secciones import (
    cargar_rangos,
    cargar_formatos,
    extraer_seccion_desde_hoja,
    generar_docx_final_en_memoria,
)
from scripts.extractor_inteligente import extraer_bloques_desde_hoja


# --- 3. CONSTANTES DE LÓGICA DE NEGOCIO ---
# Orden definido de las secciones del documento.
# Movido desde app.py para centralizar la lógica de negocio.
ORDER = [
    "Portada", "Contenido", "Dictamen 1", "Dictamen 2", "BG", "ER", "ECC", "CF",
    "Nota 1 y 2", "Nota 1 tablas", "Nota 3", "N4 Efectivo", "Nota 5 Txt",
    "N5 Inventarios Inm(Tablas)", "N6 Proveedores", "N7 Depósitos en garantía",
    "N8 Préstamos", "N9 Otras Aportaciones Fid", "Nota 10 Impuestos",
    "N11 Patrimonio", "N12 Vencimientos", "N13 Partes relacionadas",
    "Nota 14", "Nota 15", "Nota 16",
]


# --- 4. LÓGICA DE NEGOCIO CENTRALIZADA ---

def load_project_ranges() -> dict:
    """Carga los rangos de hojas desde el archivo de configuración."""
    return cargar_rangos(CONFIG_PATH)

def load_project_formats() -> dict | None:
    """Carga los formatos de hojas desde el archivo de configuración."""
    try:
        return cargar_formatos(FORMATOS_PATH)
    except FileNotFoundError:
        return None

def discover_and_load_blocks(wb: Workbook, rangos_manuales: dict, formatos_config: dict | None) -> dict:
    """
    Analiza todas las hojas de un libro de Excel y carga los bloques de contenido
    usando el método híbrido: automático primero, manual como fallback.
    """
    rangos_descubiertos = {}
    for sheet_name in wb.sheetnames:
        sheet_object = wb[sheet_name]
        
        if not formatos_config:
            formatos_config = {}

        # 1. Intentar extracción automática
        bloques_automaticos = extraer_bloques_desde_hoja(sheet_object, formatos_config)
        
        if bloques_automaticos:
            rangos_descubiertos[sheet_name] = bloques_automaticos
        elif sheet_name in rangos_manuales:
            # 2. Fallback: usar la configuración manual
            parrafos, tablas = extraer_seccion_desde_hoja(wb[sheet_name], rangos_manuales[sheet_name])
            
            bloques = []
            if parrafos:
                for p in parrafos:
                    bloques.append({'tipo': 'texto', 'contenido': p})
            if tablas:
                for t in tablas:
                    bloques.append({'tipo': 'tabla', 'contenido': t})
            
            if bloques:
                rangos_descubiertos[sheet_name] = bloques
        else:
            pass

    return rangos_descubiertos

def ejecutar_generacion_completa(
    workbook_path: str, rangos_dinamicos: dict, formatos: dict | None
) -> BytesIO:
    """
    Encapsula la generación del DOCX final, gestionando la memoria de forma explícita.
    Esta función carga el workbook, genera el documento y luego lo libera.
    """
    wb = None
    # Comentario: Se usa un bloque try...finally para garantizar que los objetos
    # pesados (el workbook) se liberen explícitamente, reduciendo la acumulación
    # de memoria en ejecuciones sucesivas de Streamlit.
    try:
        # Se vuelve a cargar el workbook aquí, pero su ciclo de vida está
        # estrictamente limitado a esta función.
        wb = load_workbook(workbook_path, data_only=True)

        buf = generar_docx_final_en_memoria(
            wb=wb,
            rangos=rangos_dinamicos,
            plantilla_path=PLANTILLA_PATH,
            orden=ORDER,
            formatos=formatos,
        )
        return buf
    finally:
        # Paso clave: Liberación explícita de memoria
        if wb:
            del wb
            # Se fuerza una recolección de basura para limpiar la memoria de inmediato.
            gc.collect()

